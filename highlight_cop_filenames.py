from __future__ import annotations

import argparse
import re
import unicodedata
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def read_lines_auto_encoding(txt_path: Path) -> list[str]:
    """Read non-empty lines with common encodings."""
    encodings = ("utf-8-sig", "gbk", "utf-16")
    last_error: Exception | None = None
    for enc in encodings:
        try:
            content = txt_path.read_text(encoding=enc)
            lines = [line.strip() for line in content.splitlines() if line.strip()]
            return lines
        except Exception as exc:  # pragma: no cover - fallback handling
            last_error = exc
    raise RuntimeError(f"无法读取文件：{txt_path}") from last_error


def normalize_filename(name: str) -> str:
    """
    Normalize text for matching while keeping version numbers effective.
    Rules:
    - keep basename if text includes path
    - normalize full/half-width chars
    - ignore spaces and decorative wrappers
    - keep dots/digits so versions still matter (e.g. v1.0.1 != v1.0.2)
    """
    n = str(name).strip().replace("\\", "/")
    if "/" in n:
        n = n.split("/")[-1]

    n = unicodedata.normalize("NFKC", n)

    # Remove common decorative quotes/brackets used in titles.
    for ch in "\"'“”‘’《》【】":
        n = n.replace(ch, "")

    # Ignore whitespace differences.
    n = re.sub(r"\s+", "", n)

    # Unify common dash variants.
    n = re.sub(r"[-_－–—]+", "-", n)

    # Lowercase for case-insensitive matching.
    n = n.lower()
    return n


def extract_filenames(lines: Iterable[str]) -> set[str]:
    names: set[str] = set()
    for line in lines:
        names.add(normalize_filename(line))
    return {n for n in names if n}


def looks_like_filename(name: str) -> bool:
    """
    Heuristic filter to avoid coloring arbitrary text.
    Example matches: a.pdf, test_v1.2.docx, xxx.tar.gz
    """
    return bool(re.search(r"\.[a-z0-9]{1,8}$", name, flags=re.IGNORECASE))


def versionless_key(name: str) -> str:
    """
    Build a filename key with version segment removed.
    Used for "same name but different version" fuzzy matching.
    """
    n = normalize_filename(name)
    if not n:
        return ""

    stem, dot, ext = n.rpartition(".")
    if not dot:
        stem, ext = n, ""

    # Remove common trailing version patterns: v1, v1.2.3, ver2.0, version3
    stem = re.sub(
        r"(?:[-_. ]*)(?:v(?:er(?:sion)?)?)[-_. ]*\d+(?:\.\d+){0,5}[a-z]?$",
        "",
        stem,
        flags=re.IGNORECASE,
    )
    # Remove numeric trailing semantic versions: -1.2, _2.0.1
    stem = re.sub(r"(?:[-_. ]*)\d+\.\d+(?:\.\d+){0,5}[a-z]?$", "", stem, flags=re.IGNORECASE)

    # Remove separators left behind after stripping version tokens.
    stem = re.sub(r"[-_. ]{2,}", "-", stem).strip("-_. ")
    if not stem:
        return ""
    return f"{stem}.{ext}" if ext else stem


def highlight_matches(txt_file: Path, xlsx_file: Path, output_file: Path) -> tuple[int, int]:
    txt_lines = read_lines_auto_encoding(txt_file)
    targets = extract_filenames(txt_lines)
    if not targets:
        raise ValueError(f"在 {txt_file} 中没有提取到可用文件名。")

    target_version_keys = {versionless_key(n) for n in targets}
    target_version_keys.discard("")

    wb = load_workbook(xlsx_file)
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFF59D")  # txt 中存在的文件名
    green_fill = PatternFill(fill_type="solid", fgColor="C8E6C9")  # 仅版本号不同

    yellow_cells = 0
    green_cells = 0

    for ws in wb.worksheets:
        max_col = ws.max_column
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max_col):
            for cell in row:
                if cell.value is None:
                    continue
                cell_text = normalize_filename(str(cell.value))
                if not looks_like_filename(cell_text):
                    continue

                if cell_text in targets:
                    cell.fill = yellow_fill
                    yellow_cells += 1
                    continue

                # Not in txt exactly, but same base name after removing version.
                if versionless_key(cell_text) in target_version_keys:
                    cell.fill = green_fill
                    green_cells += 1

    wb.save(output_file)
    return yellow_cells, green_cells


def main() -> None:
    desktop = Path.home() / "Desktop"
    parser = argparse.ArgumentParser(description="将 TXT 中文件名在 Excel 中按规则高亮")
    parser.add_argument(
        "--txt",
        type=Path,
        default=desktop / "bitmain-COP_文件名列表.txt",
        help="TXT 文件路径（默认：桌面/bitmain-COP_文件名列表.txt）",
    )
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=desktop / "bitmain知识集合-COP重复高亮.xlsx",
        help="Excel 文件路径（默认：桌面/bitmain知识集合-COP重复高亮.xlsx）",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=desktop / "bitmain知识集合-COP重复高亮_已标注.xlsx",
        help="输出 Excel 路径（默认：桌面/bitmain知识集合-COP重复高亮_已标注.xlsx）",
    )
    args = parser.parse_args()

    if not args.txt.exists():
        raise FileNotFoundError(f"找不到 TXT 文件：{args.txt}")
    if not args.xlsx.exists():
        raise FileNotFoundError(f"找不到 Excel 文件：{args.xlsx}")

    yellow_cells, green_cells = highlight_matches(args.txt, args.xlsx, args.out)
    print(f"完成：黄色高亮 {yellow_cells} 个（txt中存在）。")
    print(f"完成：绿色高亮 {green_cells} 个（仅版本号不同）。")
    print(f"输出文件：{args.out}")


if __name__ == "__main__":
    main()
