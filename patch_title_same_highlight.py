"""
将 Excel 中第二个 sheet (COP) 出现的文件名在第一个 sheet (知识集合) 中高亮标出。
匹配时忽略版本号差异，如 "文件规范-v1.3.23" 与 "文件规范-v1.0.0" 视为同一文件。
"""

import re
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

DESKTOP = Path.home() / "Desktop"
INPUT_DIR = DESKTOP / "三主体知识集合-COP"

FILES = [
    INPUT_DIR / "bitmain知识集合-COP重复高亮.xlsx",
    INPUT_DIR / "ssc知识集合-COP重复高亮.xlsx",
]

HIGHLIGHT_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

VERSION_RE = re.compile(r'[-_\s]*v\d+(\.\d+)*\s*$', re.IGNORECASE)


def strip_version(name: str) -> str:
    """去掉末尾的版本号，如 -v1.2.3 / v1.0 / _v2.1"""
    if not name:
        return ""
    return VERSION_RE.sub("", name.strip()).strip()


def process_file(filepath: Path):
    print(f"处理文件: {filepath.name}")
    wb = openpyxl.load_workbook(filepath)

    ws1 = wb[wb.sheetnames[0]]
    ws2 = wb[wb.sheetnames[1]]

    sheet2_bases = set()
    for r in range(2, ws2.max_row + 1):
        val = ws2.cell(row=r, column=1).value
        if val:
            sheet2_bases.add(strip_version(str(val)))

    highlighted = 0
    for r in range(1, ws1.max_row + 1):
        val = ws1.cell(row=r, column=1).value
        if val and strip_version(str(val)) in sheet2_bases:
            for c in range(1, ws1.max_column + 1):
                ws1.cell(row=r, column=c).fill = HIGHLIGHT_FILL
            highlighted += 1
            print(f"  高亮: 第{r}行 - {val}")

    wb.save(filepath)
    print(f"  共高亮 {highlighted} 行 (Sheet1 共 {ws1.max_row} 行)\n")


def main():
    for f in FILES:
        if f.exists():
            process_file(f)
        else:
            print(f"文件不存在: {f}")
    print("全部处理完成！")


if __name__ == "__main__":
    main()
